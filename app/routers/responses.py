"""
Responses API Endpoints
Provides CRUD operations, bulk operations, and export functionality for responses
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import io

from .. import crud, models, schemas
from ..auth import get_current_user
from ..database import SessionLocal, get_db
from ..utils.brand_access import get_active_brand_id, get_data_owner_user_id


router = APIRouter(
    prefix="/responses",
    tags=["Responses"]
)


@router.post("/", response_model=schemas.Response, status_code=201)
def create_response(
    response: schemas.ResponseCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Submit a raw response from an LLM platform for the current user's active brand."""
    return crud.create_response(db=db, response=response, user_id=current_user.id, brand_id=brand_id)


@router.get("/", response_model=List[schemas.Response])
def read_responses(
    skip: int = 0,
    limit: int = 10000,
    batch_id: Optional[int] = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id),
):
    """Retrieve responses for the active brand.

    Resolves the brand owner's user_id so that users viewing a brand that's
    been shared with them see the owner's data (Bug #1 fix). Accepts a
    `batch_id` query param to narrow to a single collection batch (Bug #2 fix).
    Default limit raised to 10000 to cover real-world brand sizes like PPPL
    (~1,200 responses); matches the precedent used by the Excel export below
    (Bug #3 fix).

    `limit` is defensively clamped to [1, 10000] so a malicious or buggy
    client can't request a negative or oversized page that would either
    return nothing or pull the entire table into memory.
    """
    limit = max(1, min(limit, 10000))
    owner_user_id = get_data_owner_user_id(db, brand_id, current_user.id)
    return crud.get_responses(
        db,
        user_id=owner_user_id,
        brand_id=brand_id,
        batch_id=batch_id,
        skip=skip,
        limit=limit,
    )


@router.get("/unanalyzed/", response_model=List[schemas.Response])
def read_unanalyzed_responses(
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Retrieve responses that are pending analysis for the current user."""
    return crud.get_unanalyzed_responses(db, user_id=current_user.id, limit=limit)


@router.put("/{response_id}/analyze", response_model=schemas.Response)
def update_response_analysis(
    response_id: int,
    analysis_data: schemas.ResponseAnalysisInput,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Update a response with analysis data for the current user's active brand."""
    db_response = crud.update_response_analysis(
        db,
        response_id=response_id,
        analysis_data=analysis_data.model_dump(exclude_unset=True),
        user_id=current_user.id,
        brand_id=brand_id
    )
    if db_response is None:
        raise HTTPException(status_code=404, detail="Response not found")
    return db_response


@router.delete("/{response_id}", response_model=schemas.Response)
def delete_response(
    response_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Delete a response for the current user's active brand."""
    deleted_response = crud.delete_response(db, response_id=response_id, user_id=current_user.id, brand_id=brand_id)
    if deleted_response is None:
        raise HTTPException(status_code=404, detail="Response not found")
    return deleted_response


@router.post("/bulk-replace-competitor")
def bulk_replace_competitor(
    old_name: str,
    new_name: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """
    Bulk find-and-replace competitor names across all responses for the active brand.

    Args:
        old_name: The competitor name to find (exact match)
        new_name: The new name to replace it with

    Returns:
        Number of responses updated
    """
    # Get all responses for the active brand
    responses = crud.get_responses(db, user_id=current_user.id, brand_id=brand_id, skip=0, limit=100000)

    updated_count = 0

    for response in responses:
        if response.competitors:
            # Split by comma, find and replace exact matches, rejoin
            competitors_list = [comp.strip() for comp in response.competitors.split(',')]

            # Replace exact matches (case-sensitive)
            updated_list = [new_name if comp == old_name else comp for comp in competitors_list]

            new_competitors = ', '.join(updated_list)

            if new_competitors != response.competitors:
                response.competitors = new_competitors
                updated_count += 1

    db.commit()

    return {"updated_count": updated_count, "old_name": old_name, "new_name": new_name}


# Excel caps a single cell at 32,767 characters. openpyxl slices silently at that
# length (openpyxl/cell/cell.py), so a long grounded answer loses its tail with no
# error and no marker in the file. We truncate deliberately instead, leaving a
# visible note that points at the CSV export, which has no such limit.
EXCEL_CELL_LIMIT = 32767
_TRUNCATION_NOTE = "  [truncated at {limit} characters, the Excel cell limit; use the CSV export for the full text]"


def _excel_safe(value: Optional[str]) -> str:
    """Make a response body safe to put in a worksheet cell.

    Two failure modes, both of which currently break the export:

    1. Control characters. openpyxl raises IllegalCharacterError when the value is
       assigned, before the workbook is ever saved. Because the export is
       all-or-nothing, one stray byte anywhere in a brand's data made the whole
       spreadsheet unavailable. LLM answers do occasionally carry stray control
       bytes, so they are stripped here rather than allowed to fail the request.
    2. Length. See EXCEL_CELL_LIMIT above.
    """
    if not value:
        return ''
    # Strip the characters in openpyxl's ILLEGAL_CHARACTERS_RE, keeping tab,
    # newline and carriage return, which Excel accepts.
    cleaned = ''.join(
        ch for ch in value
        if ch in '\t\n\r' or ord(ch) >= 32
    )
    if len(cleaned) > EXCEL_CELL_LIMIT:
        note = _TRUNCATION_NOTE.format(limit=EXCEL_CELL_LIMIT)
        cleaned = cleaned[:EXCEL_CELL_LIMIT - len(note)] + note
    return cleaned


@router.get("/export/excel")
def export_responses_to_excel(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand_id: Optional[int] = Depends(get_active_brand_id)
):
    """Export all responses for the active brand to an Excel spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    # Resolve the brand OWNER, not the requesting user. Every other read path on
    # this router does the same (see read_responses); using current_user.id here
    # meant a user viewing a shared brand saw the table on screen but got a 404
    # from the export button.
    owner_user_id = get_data_owner_user_id(db, brand_id, current_user.id)

    # Page through rather than capping at a fixed limit. The previous limit=10000
    # silently dropped every row beyond it, so a large brand's spreadsheet was
    # quietly incomplete.
    responses = []
    page_size = 5000
    while True:
        page = crud.get_responses(
            db, user_id=owner_user_id, brand_id=brand_id,
            skip=len(responses), limit=page_size,
        )
        responses.extend(page)
        if len(page) < page_size:
            break

    if not responses:
        raise HTTPException(status_code=404, detail="No responses found for export")

    # Get brand name for filename
    brand = crud.get_brand_by_id(db, brand_id, owner_user_id) if brand_id else None
    brand_name = brand.brand_name if brand else "Unknown"

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Responses"

    # Define headers
    headers = [
        "Response ID",
        # Batch ID is what the dashboard's batch selector scopes by. Without it
        # there is no way to reconcile a row in this spreadsheet against a number
        # on screen.
        "Batch ID",
        "Query ID",
        # The exported row previously identified its question only by id, so you
        # could not tell what "Q014" had actually asked.
        "Query Text",
        "Platform",
        "Response Text",
        "Collected At (UTC)",
        "Analyzed At (UTC)",
        "Brand Mentioned",
        "Brand Position",
        "Sentiment",
        "Descriptors",
        "Competitors",
        "Sources",
        "Notes"
    ]

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for row_num, response in enumerate(responses, 2):
        ws.cell(row=row_num, column=1).value = response.id
        ws.cell(row=row_num, column=2).value = response.batch_id
        ws.cell(row=row_num, column=3).value = response.query_id
        ws.cell(row=row_num, column=4).value = _excel_safe(response.query_text)
        ws.cell(row=row_num, column=5).value = response.platform
        ws.cell(row=row_num, column=6).value = _excel_safe(response.response_text)
        ws.cell(row=row_num, column=7).value = response.timestamp.strftime('%Y-%m-%d %H:%M:%S') if response.timestamp else ''
        ws.cell(row=row_num, column=8).value = response.analyzed_at.strftime('%Y-%m-%d %H:%M:%S') if response.analyzed_at else ''
        ws.cell(row=row_num, column=9).value = _excel_safe(response.brand_mentioned)
        ws.cell(row=row_num, column=10).value = _excel_safe(response.brand_position)
        ws.cell(row=row_num, column=11).value = _excel_safe(response.sentiment)
        ws.cell(row=row_num, column=12).value = _excel_safe(response.descriptors)
        ws.cell(row=row_num, column=13).value = _excel_safe(response.competitors)
        ws.cell(row=row_num, column=14).value = _excel_safe(response.sources)
        ws.cell(row=row_num, column=15).value = _excel_safe(response.notes)

    # Auto-adjust column widths
    for col in ws.columns:
        column = col[0].column_letter
        max_length = max((len(str(cell.value)) for cell in col), default=0)
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Create safe filename
    safe_brand_name = "".join(c for c in brand_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f"{safe_brand_name}_AI_Responses.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
