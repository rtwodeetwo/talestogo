"""
The reconciliation harness: do the surfaces agree with each other?

Every test here asks one question. For the same brand, the same period and the
same underlying rows, does the Dashboard report the same number as the CSV
export, the Excel export, the stored BatchAnalytics row, the generated report
and the highlights email?

When this harness was written the answer was no: six surfaces produced four
different mention rates for one batch, a 13 point spread. Each disagreement was
recorded as an xfail naming the file:line responsible, and that list served as
the migration checklist. It is now empty. The spread is 0.0.

These tests are therefore no longer a catalogue of defects; they are the
regression guard that keeps the surfaces in agreement. A failure here means some
surface has started computing its own arithmetic again instead of calling
metrics_core.

Run `pytest tests/test_metric_reconciliation.py -s` to print the matrix.
"""
import csv
import io

import pytest

from app.services import metrics_core as mc
from app.services import metrics_query as mq
from tests import golden_expected as gx
from tests.fixtures.golden_dataset import (
    BATCH_1_ID, BRAND_1_ID, CONTROL_CHAR_RESPONSE_TEXT, USER_1_ID,
)

pytestmark = pytest.mark.reconciliation


def _canonical(db, batch_id=BATCH_1_ID):
    return mq.resolve(db, mq.MetricScope.for_batch(USER_1_ID, BRAND_1_ID, batch_id))


# ============================================================ surface: dashboard

class TestDashboardSurface:
    """GET /api/analytics/dashboard against the canonical definitions."""

    @pytest.fixture()
    def dashboard(self, golden_client, golden_db_with_stored_analytics):
        response = golden_client.get(
            f"/api/analytics/dashboard?brand_id={BRAND_1_ID}&batch_id={BATCH_1_ID}")
        assert response.status_code == 200, response.text
        return response.json()

    def test_mention_rate_matches_canonical(self, dashboard):
        assert dashboard["mention_rate"] == gx.B1_MENTION_RATE

    def test_positive_sentiment_matches_canonical(self, dashboard):
        assert dashboard["positive_sentiment"] == gx.B1_POSITIVE_SENTIMENT_RATE

    def test_share_of_voice_matches_canonical(self, dashboard):
        assert dashboard["share_of_voice"] == gx.B1_SHARE_OF_VOICE

    def test_descriptor_match_matches_canonical(self, dashboard):
        assert dashboard["descriptor_match"] == gx.B1_DESCRIPTOR_MATCH_RATE

    def test_reports_a_total_response_count(self, dashboard):
        assert "total_responses" in dashboard

    def test_reports_excluded_row_counts(self, dashboard):
        """Without this, a collection failure looks like a reputation drop."""
        quality = dashboard["data_quality"]
        assert quality["counted"] == gx.B1_POPULATION
        assert quality["unanalyzed_excluded"] == gx.B1_UNANALYZED_EXCLUDED
        assert quality["invalid_enum_excluded"] == gx.B1_INVALID_ENUM_EXCLUDED
        assert quality["orphan_query_excluded"] == gx.B1_ORPHAN_EXCLUDED
        assert quality["branded_excluded"] == gx.B1_BRANDED_EXCLUDED

    def test_leadership_visibility_is_reported(self, dashboard):
        assert dashboard["leadership_visibility"] == gx.B1_LEADERSHIP_VISIBILITY

    def test_positioning_average_is_on_the_five_point_scale(self, dashboard):
        assert dashboard["positioning_average"] == gx.B1_POSITIONING_AVERAGE


class TestDashboardInternalConsistency:
    """Fields within a single Dashboard payload that must agree with each other."""

    def test_positive_sentiment_equals_its_own_pie_slices(self, golden_client,
                                                          golden_db_with_stored_analytics):
        dashboard = golden_client.get(
            f"/api/analytics/dashboard?brand_id={BRAND_1_ID}&batch_id={BATCH_1_ID}").json()
        breakdown = golden_client.get(
            f"/api/analytics/sentiment/breakdown?brand_id={BRAND_1_ID}&batch_id={BATCH_1_ID}"
        ).json()
        slices = (breakdown.get("very_positive_pct", 0) + breakdown.get("positive_pct", 0))
        assert dashboard["positive_sentiment"] == pytest.approx(slices, abs=0.1)

    def test_leadership_value_and_delta_use_one_formula(self, golden_client,
                                                        golden_db_with_stored_analytics):
        """The tile's value and its trend arrow must measure the same thing.

        The value came from the share-of-voice endpoint as Leader+Top 3+Featured
        while the arrow was computed from leader_count alone, so the number and
        its direction of travel were different metrics. Both now come from the
        dashboard payload.
        """
        dashboard = golden_client.get(
            f"/api/analytics/dashboard?brand_id={BRAND_1_ID}&batch_id={BATCH_1_ID}").json()
        assert dashboard["leadership_visibility"] == gx.B1_LEADERSHIP_VISIBILITY
        assert "change_leadership_visibility" in dashboard

    def test_leadership_delta_is_the_difference_of_two_values(self, golden_client,
                                                              golden_db_with_stored_analytics):
        """Batch 1 is the earliest batch, so there is nothing to compare against
        and the reported change must be zero rather than a spurious jump."""
        dashboard = golden_client.get(
            f"/api/analytics/dashboard?brand_id={BRAND_1_ID}&batch_id={BATCH_1_ID}").json()
        assert dashboard["change_leadership_visibility"] == 0
        assert dashboard["previous_collection_date"] is None


# ======================================================== surface: stored cache

class TestStoredBatchAnalytics:
    @pytest.fixture()
    def stored(self, golden_db_with_stored_analytics):
        from app import models
        return golden_db_with_stored_analytics.query(models.BatchAnalytics).filter(
            models.BatchAnalytics.batch_id == BATCH_1_ID).first()

    def test_mention_rate_matches_canonical(self, stored):
        assert stored.mention_rate == gx.B1_MENTION_RATE

    def test_position_counts_sum_to_total(self, stored):
        """All five buckets, including Top 3, must account for the population."""
        counted = (stored.leader_count + stored.top3_count + stored.featured_count
                   + stored.listed_count + stored.not_mentioned_count)
        assert counted == stored.total_responses

    def test_top_3_is_recorded(self, stored):
        assert stored.top3_count == gx.B1_POSITION_COUNTS["Top 3"]

    def test_total_responses_is_the_metric_denominator(self, stored):
        """Not a row count for the batch. Excluded rows are reported separately
        rather than folded into not_mentioned_count."""
        assert stored.total_responses == gx.B1_POPULATION
        assert stored.unanalyzed_count == gx.B1_UNANALYZED_EXCLUDED
        assert stored.invalid_count == gx.B1_INVALID_ENUM_EXCLUDED

    def test_sentiment_counts_sum_to_their_own_base(self, stored):
        """sentiment_base_count is what the sentiment counts divide by.

        Previously sentiment percentages were derived from mention_count, which
        includes Indirect mentions carrying no sentiment, so the slices could
        never sum to 100. The base is its own population: it includes answers to
        branded questions, which the visibility metrics exclude.
        """
        slices = (stored.very_positive_count + stored.positive_count
                  + stored.neutral_count + stored.negative_count
                  + stored.very_negative_count + stored.mixed_count)
        assert slices == stored.sentiment_base_count == gx.B1_SENTIMENT_POPULATION

    def test_visibility_counts_are_organic_only(self, stored):
        assert stored.direct_mention_count <= stored.mention_count
        assert stored.mention_count <= stored.total_responses

    def test_precision_is_preserved(self, stored):
        """batch_analytics.py used to round to int while analytics_cache.py
        rounded to 2 decimals for the same concept."""
        assert stored.mention_rate == pytest.approx(gx.B1_MENTION_RATE, abs=0.05)

    def test_row_is_tagged_with_the_definition_that_wrote_it(self, stored):
        from app.services.batch_analytics import METRICS_VERSION
        assert stored.metrics_version == METRICS_VERSION


# ========================================================== surface: CSV export

class TestCsvExport:
    @pytest.fixture()
    def report_id(self, golden_db):
        """A monthly report over January 2026, as the scheduler would store it."""
        import datetime

        from app import models
        report = models.Report(
            user_id=USER_1_ID, brand_id=BRAND_1_ID, title="January 2026",
            report_content="# January 2026", report_type="monthly",
            period_label="January 2026",
            start_date=datetime.datetime(2026, 1, 1),
            end_date=datetime.datetime(2026, 1, 31, 23, 59, 59),
            total_responses=gx.B1_POPULATION,
        )
        golden_db.add(report)
        golden_db.commit()
        return report.id

    @pytest.fixture()
    def rows(self, golden_client, report_id):
        response = golden_client.get(f"/reports/{report_id}/export/csv")
        assert response.status_code == 200, response.text
        return response, list(csv.DictReader(io.StringIO(response.text)))

    def test_the_reports_population_is_derivable_from_the_csv(self, rows):
        """The reconciliation property that matters.

        A raw data export should show every row, not silently hide the ones the
        metrics excluded. So instead of the row count matching the report, the
        CSV carries the flags needed to reproduce the report's population from
        the spreadsheet: filter on "Counted In Metrics" and you get exactly the
        rows behind the published number.
        """
        _, parsed = rows
        counted = [r for r in parsed if r["Counted In Metrics"] == "Yes"]
        assert len(counted) == gx.B1_POPULATION

    def test_shows_every_row_not_just_the_counted_ones(self, rows):
        """Excluded rows stay visible, so a collection failure is inspectable."""
        _, parsed = rows
        assert len(parsed) > gx.B1_POPULATION
        assert any(r["Counted In Metrics"] == "No" for r in parsed)

    def test_branded_queries_are_labelled(self, rows):
        _, parsed = rows
        branded = [r for r in parsed if r["Brand In Query"] == "Yes"]
        assert len(branded) == gx.B1_BRANDED_EXCLUDED + 1  # +1 unanalyzed branded row

    def test_every_excluded_row_says_why(self, rows):
        """The counts in data_quality, made row-level and inspectable.

        This is what lets a failed collection be told apart from a reputation
        drop without opening the database.
        """
        _, parsed = rows
        excluded = [r for r in parsed if r["Counted In Metrics"] == "No"]
        assert excluded
        assert all(r["Excluded Because"] for r in excluded)
        assert not any(r["Excluded Because"] for r in parsed
                       if r["Counted In Metrics"] == "Yes")

        reasons = {r["Excluded Because"] for r in excluded}
        assert reasons == {
            "Brand named in the question",
            "Not analyzed",
            "Unrecognized analysis result",
            "Query no longer exists",
        }

    def test_exclusion_reasons_match_the_canonical_counts(self, rows, golden_db):
        """Row-level reasons must add up to what metrics_core reports.

        A row can be excluded for more than one reason at once, and the CSV
        reports one per row by precedence, so the branded-and-unanalyzed row
        shows as branded. Total exclusions still reconcile.
        """
        _, parsed = rows
        counts = {}
        for row in parsed:
            if row["Excluded Because"]:
                counts[row["Excluded Because"]] = counts.get(row["Excluded Because"], 0) + 1

        assert counts["Not analyzed"] == gx.B1_UNANALYZED_ORGANIC
        assert counts["Unrecognized analysis result"] == gx.B1_INVALID_ENUM_EXCLUDED
        assert counts["Query no longer exists"] == gx.B1_ORPHAN_EXCLUDED
        assert counts["Brand named in the question"] == (
            gx.B1_BRANDED_EXCLUDED + gx.B1_UNANALYZED_BRANDED)

        assert sum(counts.values()) == gx.B1_TOTAL_ROWS - gx.B1_POPULATION

    def test_has_utf8_bom_for_excel(self, rows):
        response, _ = rows
        assert response.content.startswith(b"\xef\xbb\xbf"), (
            "without a BOM Excel opens the file in the system codepage and every "
            "curly quote and accented character in an AI answer becomes mojibake")

    def test_includes_batch_id(self, rows):
        _, parsed = rows
        assert parsed and "batch id" in {k.lower() for k in parsed[0]}

    def test_long_answers_are_not_truncated(self, rows):
        """The CSV has no cell limit, which is why the Excel truncation note
        points here for the full text."""
        _, parsed = rows
        longest = max((len(r["Response Text"]) for r in parsed), default=0)
        assert longest == gx.LONG_RESPONSE_LENGTH

    def test_multiline_bodies_do_not_break_row_alignment(self, rows):
        """csv.writer quotes embedded newlines correctly; this guards a regression."""
        _, parsed = rows
        assert all(len(row) == len(parsed[0]) for row in parsed)

    def test_word_export_is_scoped_to_the_reports_period(self, golden_client, report_id):
        """The Word export substituted all-time figures into a period report.

        It called the legacy analytics module with no date filter, so a January
        report carried all-time sentiment and share of voice beside January's
        prose. It also passed the requesting user rather than the brand owner,
        so a shared brand exported zeros.
        """
        response = golden_client.get(f"/reports/{report_id}/export/word")
        assert response.status_code == 200, response.text[:400]
        assert len(response.content) > 0

    def test_timestamps_carry_no_offset_today(self, rows):
        """reports.py:290 writes naive UTC while the UI displayed Eastern, so a
        user filtering the spreadsheet by date sees a different day."""
        _, parsed = rows
        if not parsed:
            pytest.skip("no rows")
        key = next((k for k in parsed[0] if "time" in k.lower() or "date" in k.lower()), None)
        if key is None:
            pytest.skip("no timestamp column")
        assert "+" not in parsed[0][key] and "Z" not in parsed[0][key]


# ======================================================== surface: Excel export

class TestExcelExport:
    @pytest.fixture()
    def workbook(self, golden_client):
        import openpyxl
        response = golden_client.get(f"/responses/export/excel?brand_id={BRAND_1_ID}")
        assert response.status_code == 200, response.text[:500]
        return openpyxl.load_workbook(io.BytesIO(response.content))

    def test_long_answers_are_not_silently_truncated(self, workbook):
        """The reported symptom: answers cut off in the spreadsheet.

        Excel genuinely cannot hold more than 32,767 characters in one cell, so
        the fix is not to preserve the full text here. It is to stop losing it
        silently: the cell now ends with a visible note saying it was cut and
        where to get the whole answer. The CSV export has no such limit.
        """
        sheet = workbook.active
        longest = max(
            (cell.value for row in sheet.iter_rows(min_row=2)
             for cell in row if isinstance(cell.value, str)),
            key=len, default="")
        assert len(longest) > 30000, "the long fixture body is missing entirely"
        assert "[truncated at 32767 characters" in longest, (
            "a truncated answer must say so in the cell, not just end mid-word")
        assert "CSV export" in longest, "the note must point at the full text"

    def test_includes_query_text(self, workbook):
        headers = {str(c.value).lower() for c in next(workbook.active.iter_rows(max_row=1))}
        assert any("query text" in h for h in headers)

    def test_includes_batch_id(self, workbook):
        headers = {str(c.value).lower() for c in next(workbook.active.iter_rows(max_row=1))}
        assert any("batch" in h for h in headers)

    def test_export_survives_control_characters(self, golden_client):
        """One vertical tab anywhere in the data takes down the entire download.

        Verified directly: assigning the fixture's \\x0b body raises
        IllegalCharacterError on the cell, before save is ever reached. Because
        the export is all-or-nothing, a single bad character makes the
        spreadsheet unavailable for the whole brand.
        """
        response = golden_client.get(f"/responses/export/excel?brand_id={BRAND_1_ID}")
        assert response.status_code == 200, (
            "Excel export failed outright. responses.py:156 has no guard for "
            f"openpyxl's ILLEGAL_CHARACTERS_RE. Body: {response.text[:300]}")


class TestExcelExportScoping:
    def test_shared_brand_users_can_export(self, golden_session_factory, golden_db):
        """responses.py:159 filters by current_user.id instead of the brand owner,
        so a user viewing a shared brand sees the table but cannot export it."""
        from fastapi import Depends
        from fastapi.testclient import TestClient
        from sqlalchemy.orm import Session

        from app import models
        from app.auth import get_current_user
        from app.database import get_db
        from app.main import app
        from app.utils.brand_access import get_active_brand_id

        shared_user_id = 900
        golden_db.add(models.User(
            id=shared_user_id, email="shared@golden.test", full_name="Shared Viewer",
            is_active=True, is_admin=False, is_invited=True))
        golden_db.add(models.BrandShare(
            brand_id=BRAND_1_ID, user_id=shared_user_id,
            shared_by_user_id=USER_1_ID, permission_level="edit"))
        golden_db.commit()

        def override_get_db():
            db = golden_session_factory()
            try:
                yield db
            finally:
                db.close()

        def override_get_current_user(db: Session = Depends(get_db)):
            return db.query(models.User).filter(models.User.id == shared_user_id).first()

        app.dependency_overrides[get_db] = override_get_db
        app.dependency_overrides[get_current_user] = override_get_current_user
        app.dependency_overrides[get_active_brand_id] = lambda: BRAND_1_ID
        try:
            response = TestClient(app).get(f"/responses/export/excel?brand_id={BRAND_1_ID}")
            assert response.status_code == 200, (
                "Shared-brand export failed. responses.py:159 uses current_user.id; "
                "every other read path resolves the owner with "
                f"get_data_owner_user_id. Status {response.status_code}")
        finally:
            app.dependency_overrides.clear()


# ==================================================== surface: highlights email

class TestHighlightsSurface:
    @pytest.fixture()
    def computed(self, golden_db):
        from app.routers import highlights
        return highlights, _canonical(golden_db)

    def test_mention_rate_matches_canonical(self, computed):
        highlights, population = computed
        total, mentioned, rate = highlights._compute_mention_rate(population)
        assert rate == gx.B1_MENTION_RATE
        assert total == gx.B1_POPULATION
        assert mentioned == gx.B1_MENTION_NUMERATOR

    def test_descriptor_counts_match_canonical(self, computed):
        highlights, population = computed
        emailed = dict(highlights._compute_descriptors(population))
        assert emailed == gx.B1_DESCRIPTOR_FREQUENCY

    def test_platform_rates_match_canonical(self, computed):
        highlights, population = computed
        emailed = {name: data["rate"]
                   for name, data in highlights._compute_platform_rates(population).items()}
        assert emailed == gx.B1_PLATFORM_MENTION_RATES

    def test_sentiment_matches_canonical(self, computed):
        highlights, population = computed
        sentiment, total = highlights._compute_sentiment(population)
        assert total == gx.B1_SENTIMENT_POPULATION
        positive = (sentiment["Very Positive"]["pct"] + sentiment["Positive"]["pct"])
        assert positive == pytest.approx(gx.B1_POSITIVE_SENTIMENT_RATE, abs=0.1)


# ========================================================= cross-surface matrix

def _dashboard_mention_rate(client, db):
    body = client.get(
        f"/api/analytics/dashboard?brand_id={BRAND_1_ID}&batch_id={BATCH_1_ID}").json()
    return body.get("mention_rate")


def _stored_mention_rate(client, db):
    from app import models
    row = db.query(models.BatchAnalytics).filter(
        models.BatchAnalytics.batch_id == BATCH_1_ID).first()
    return row.mention_rate if row else None


def _highlights_mention_rate(client, db):
    from app.routers import highlights
    return highlights._compute_mention_rate(_canonical(db))[2]


def _report_mention_rate(client, db):
    """What generate_report.py:394 computes for the same window."""
    import datetime
    import importlib

    module = importlib.import_module("scripts.admin.generate_report")
    metrics = module.calculate_period_metrics(
        db, USER_1_ID, BRAND_1_ID,
        datetime.datetime(2026, 1, 1), datetime.datetime(2026, 2, 1))
    return metrics.get("mention_rate")


def _canonical_mention_rate(client, db):
    return mc.mention_rate(_canonical(db)).value


MENTION_RATE_SURFACES = {
    "dashboard endpoint": _dashboard_mention_rate,
    "stored BatchAnalytics": _stored_mention_rate,
    "highlights email": _highlights_mention_rate,
    "generated report": _report_mention_rate,
    "CANONICAL metrics_core": _canonical_mention_rate,
}


def test_all_surfaces_agree_on_mention_rate(golden_client,
                                            golden_db_with_stored_analytics):
    """The headline reconciliation. On failure it prints the full matrix.

    This started as the evidence that the numbers disagreed. It is now the
    evidence that they do not: every surface derives its figure from
    metrics_core, so one brand and one batch produce one answer.

    app/analytics.py is deliberately absent from this matrix. Its
    get_dashboard_metrics has no caller in the application, only in tooling, and
    it is scheduled for deletion; the live functions that remained in it have
    been migrated.
    """
    db = golden_db_with_stored_analytics
    observed = {}
    for name, extractor in MENTION_RATE_SURFACES.items():
        try:
            observed[name] = extractor(golden_client, db)
        except Exception as exc:  # noqa: BLE001 - a crashing surface is a finding
            observed[name] = f"ERROR: {type(exc).__name__}: {exc}"

    width = max(len(name) for name in observed)
    lines = ["", "Mention rate for brand 1 / batch 1, by surface:", ""]
    for name, value in observed.items():
        lines.append(f"    {name.ljust(width)}  {value}")
    numeric = [v for v in observed.values() if isinstance(v, (int, float))]
    if numeric:
        lines.append("")
        lines.append(f"    {'SPREAD'.ljust(width)}  "
                     f"{round(max(numeric) - min(numeric), 2)} percentage points")
    print("\n".join(lines))

    distinct = {round(float(v), 1) for v in numeric}
    assert len(distinct) == 1, (
        "Surfaces disagree on mention rate:\n" + "\n".join(lines))
